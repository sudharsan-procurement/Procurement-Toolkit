"""
Stage 5 (partial) — EXPORT.

For the MVP we export the ranked change list to an Excel (.xlsx) file. Annotated
PDF / Word output is on the roadmap for later; a spreadsheet is the most useful
first format because reviewers can sort, filter, and sign off row by row.
"""

from __future__ import annotations

import io

from .compare import Change


def changes_to_excel(changes: list[Change]) -> bytes:
    """Return an .xlsx file (as bytes) containing one row per change."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Changes"

    headers = [
        "Rank", "Clause", "Category", "Severity",
        "Numbers changed", "Old text", "New text",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for rank, c in enumerate(changes, start=1):
        nums = "; ".join(
            f"{nc.old} -> {nc.new} ({nc.description})" for nc in c.number_changes
        )
        ws.append([
            rank, c.label, c.category, c.severity,
            nums, c.old_text, c.new_text,
        ])

    # Reasonable column widths and wrapping for the long text columns.
    widths = {"A": 6, "B": 12, "C": 16, "D": 9, "E": 40, "F": 60, "G": 60}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
