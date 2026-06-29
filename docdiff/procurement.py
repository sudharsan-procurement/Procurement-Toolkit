"""
Procurement Toolkit — a SEPARATE module of procurement-specific tools.

    compare_quotes / build_quote_excel        -> multi-vendor quote comparison
    validate_po_invoice / build_validation_excel -> PO vs invoice field validation

It reuses docdiff.tables.file_to_dataframe() to load CSV/Excel/Word/PDF/image
tables into DataFrames, and follows the same openpyxl export pattern as the rest
of the app. Nothing here touches the existing document tools.
"""

from __future__ import annotations

import io
import re

from .tables import file_to_dataframe


# --- Shared helpers -----------------------------------------------------------
def load_table(file_bytes: bytes, filename: str):
    """Load a procurement file into a DataFrame."""
    return file_to_dataframe(file_bytes, filename)


def _to_num(value):
    """Parse a price/quantity, stripping currency symbols, commas, spaces."""
    if value is None:
        return None
    s = re.sub(r"[^\d.\-]", "", str(value))
    if s in ("", "-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


# --- Feature 1: Quote Comparison ---------------------------------------------
def _unique_descs(vendors: list[dict]) -> list[str]:
    """Every distinct item description across all vendors, in first-seen order."""
    seen: list[str] = []
    found: set[str] = set()
    for v in vendors:
        for _, row in v["df"].iterrows():
            d = str(row.get(v["item_col"], "")).strip()
            if d and d.lower() != "nan" and d not in found:
                found.add(d)
                seen.append(d)
    return seen


def _cluster_difflib(descs: list[str], threshold: float) -> dict:
    """Fallback fuzzy matcher (no model): group by character-similarity ratio."""
    from difflib import SequenceMatcher

    reps: list[str] = []
    mapping: dict[str, str] = {}
    for d in descs:
        best, best_ratio = None, 0.0
        for r in reps:
            ratio = SequenceMatcher(None, d.lower(), r.lower()).ratio()
            if ratio > best_ratio:
                best, best_ratio = r, ratio
        if best is not None and best_ratio >= threshold:
            mapping[d] = best
        else:
            reps.append(d)
            mapping[d] = d
    return mapping


def cluster_descriptions(descs: list[str], threshold: float, model=None) -> dict:
    """
    Group similar item descriptions together by MEANING, returning a mapping of
    each description -> its group's representative description.

    Reuses the Compare tool's local embedding model (sentence-transformers). If
    the model isn't available, falls back to character-similarity matching.
    """
    if not descs:
        return {}
    if model is None:
        from .align import _load_model
        model = _load_model()
    if model is None:
        return _cluster_difflib(descs, threshold)

    import numpy as np

    embs = model.encode(descs, convert_to_numpy=True, normalize_embeddings=True)
    rep_idx: list[int] = []
    rep_embs: list = []
    mapping: dict[str, str] = {}
    for i, d in enumerate(descs):
        if rep_embs:
            cos = np.asarray(rep_embs) @ embs[i]   # cosine (vectors are normalized)
            sims = (cos + 1.0) / 2.0               # scale to 0..1, like align.py
            j = int(sims.argmax())
            if sims[j] >= threshold:
                mapping[d] = descs[rep_idx[j]]
                continue
        rep_idx.append(i)
        rep_embs.append(embs[i])
        mapping[d] = d
    return mapping


def compare_quotes(vendors: list[dict], match_mode: str = "exact",
                   fuzzy_threshold: float = 0.75, model=None) -> dict:
    """
    Compare quotes from several vendors, matching rows by item description.

    `vendors`: list of {name, df, item_col, price_col, qty_col(optional)}.
    `match_mode`: "exact" (normalized text) or "fuzzy" (meaning-based grouping).
    Returns a structured result used for both the on-screen view and the Excel.
    """
    names = [v["name"] for v in vendors]
    items: dict[str, dict] = {}
    order: list[str] = []

    # In fuzzy mode, pre-cluster all distinct descriptions into groups.
    group_of = (cluster_descriptions(_unique_descs(vendors), fuzzy_threshold, model)
                if match_mode == "fuzzy" else None)

    for v in vendors:
        df, ic, pc, qc = v["df"], v["item_col"], v["price_col"], v.get("qty_col")
        for _, row in df.iterrows():
            desc = str(row.get(ic, "")).strip()
            if not desc or desc.lower() == "nan":
                continue
            rep = group_of.get(desc, desc) if group_of else desc
            key = _norm_text(rep)
            if key not in items:
                items[key] = {"desc": rep, "qty": None, "prices": {}, "aliases": set()}
                order.append(key)
            items[key]["aliases"].add(desc)
            price = _to_num(row.get(pc))
            if price is not None:
                # If a vendor's wording matched this group more than once, keep the
                # lowest price they offered for it.
                prev = items[key]["prices"].get(v["name"])
                items[key]["prices"][v["name"]] = price if prev is None else min(prev, price)
            if qc:
                qv = _to_num(row.get(qc))
                if qv is not None and items[key]["qty"] is None:
                    items[key]["qty"] = qv

    rows = []
    vendor_totals = {n: 0.0 for n in names}
    vendor_quoted = {n: 0 for n in names}
    vendor_price_sum = {n: 0.0 for n in names}
    items_won = {n: 0 for n in names}
    optimal_total = 0.0

    for key in order:
        it = items[key]
        prices = it["prices"]
        qty = it["qty"] or 1.0
        valid = {n: p for n, p in prices.items() if p is not None}

        row = {"Item": it["desc"], "Quantity": qty,
               "_aliases": sorted(it.get("aliases", []))}
        for n in names:
            row[n] = prices.get(n)  # None where a vendor didn't quote the item

        if valid:
            low_v = min(valid, key=valid.get)
            low_p = valid[low_v]
            high_p = max(valid.values())
            items_won[low_v] += 1
            optimal_total += low_p * qty
            row["Lowest Vendor"] = low_v
            row["Lowest Price"] = low_p
            row["_high_price"] = high_p
        else:
            row["Lowest Vendor"] = None
            row["Lowest Price"] = None
            row["_high_price"] = None
        rows.append(row)

        for n, p in valid.items():
            vendor_totals[n] += p * qty
            vendor_price_sum[n] += p
            vendor_quoted[n] += 1

    averages = {
        n: (vendor_price_sum[n] / vendor_quoted[n] if vendor_quoted[n] else 0.0)
        for n in names
    }
    # "Overall" totals are only comparable between vendors who quoted EVERY item;
    # an incomplete quote looks artificially cheap. Fall back to all quoting
    # vendors only if nobody quoted the full list.
    n_items = len(order)
    complete = [n for n in names if vendor_quoted[n] == n_items and n_items > 0]
    incomplete = [n for n in names if 0 < vendor_quoted[n] < n_items]
    basis = complete if complete else [n for n in names if vendor_quoted[n] > 0]

    lowest_overall = min(basis, key=lambda n: vendor_totals[n]) if basis else None
    highest_overall = max(basis, key=lambda n: vendor_totals[n]) if basis else None
    # Savings are measured against the optimal cherry-picked basket and clamped
    # at 0 (negative would only mean an unfair incomplete comparison).
    potential_savings = (
        max(0.0, vendor_totals[lowest_overall] - optimal_total) if lowest_overall else 0.0
    )
    savings_opportunity = (
        max(0.0, vendor_totals[highest_overall] - optimal_total) if highest_overall else 0.0
    )

    return {
        "vendors": names,
        "rows": rows,
        "vendor_totals": vendor_totals,
        "items_won": items_won,
        "averages": averages,
        "optimal_total": optimal_total,
        "lowest_overall": lowest_overall,
        "highest_overall": highest_overall,
        "potential_savings": potential_savings,
        "savings_opportunity": savings_opportunity,
        "total_items": n_items,
        "incomplete_vendors": incomplete,
        "complete_basis": bool(complete),
    }


def build_quote_excel(result: dict) -> bytes:
    """Excel report: Summary / Item Comparison (highlighted) / Savings Analysis."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    names = result["vendors"]
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HEAD = PatternFill("solid", fgColor="1F2937")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Total items compared", result["total_items"]])
    ws.append(["Lowest overall vendor", result["lowest_overall"] or "—"])
    ws.append(["Optimal basket (cheapest per item)", round(result["optimal_total"], 2)])
    ws.append(["Potential savings (vs cheapest single vendor)",
               round(result["potential_savings"], 2)])
    ws.append(["Savings opportunity (vs most expensive vendor)",
               round(result["savings_opportunity"], 2)])
    ws.append([])
    ws.append(["Vendor", "Total quote value", "Items won", "Average item price"])
    for n in names:
        ws.append([n, round(result["vendor_totals"][n], 2),
                   result["items_won"][n], round(result["averages"][n], 2)])

    # Sheet 2 — Item Comparison with green/red/yellow highlighting
    ws2 = wb.create_sheet("Item Comparison")
    ws2.append(["Item", "Quantity"] + names + ["Lowest Vendor", "Lowest Price"])
    for r in result["rows"]:
        ws2.append([r["Item"], r["Quantity"]] + [r.get(n) for n in names]
                   + [r.get("Lowest Vendor"), r.get("Lowest Price")])
    for ri, r in enumerate(result["rows"], start=2):
        low, high = r.get("Lowest Price"), r.get("_high_price")
        for ci, n in enumerate(names):
            cell = ws2.cell(row=ri, column=3 + ci)  # vendors start at col 3
            val = r.get(n)
            if val is None:
                cell.fill = YELLOW
            elif low is not None and val == low:
                cell.fill = GREEN
            elif high is not None and val == high and high != low:
                cell.fill = RED

    # Sheet 3 — Savings Analysis
    ws3 = wb.create_sheet("Savings Analysis")
    ws3.append(["Item", "Quantity", "Lowest Vendor", "Lowest Price",
                "Highest Price", "Saving per unit", "Saving total"])
    for r in result["rows"]:
        low, high = r.get("Lowest Price"), r.get("_high_price")
        qty = r.get("Quantity") or 1
        if low is not None and high is not None:
            ws3.append([r["Item"], qty, r.get("Lowest Vendor"), low, high,
                        round(high - low, 2), round((high - low) * qty, 2)])

    for sheet in (ws, ws2, ws3):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEAD

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Feature 2: PO vs Invoice Validator --------------------------------------
PO_FIELDS = [
    "Vendor Name", "GST Number", "Item Description",
    "Quantity", "Unit Price", "Tax Amount", "Total Amount",
]
_SUM_FIELDS = {"Quantity", "Tax Amount", "Total Amount"}
_NUMERIC_FIELDS = {"Quantity", "Unit Price", "Tax Amount", "Total Amount"}


def _extract_field(df, col, field):
    """Reduce a mapped column to a single comparable value for the field.

    Amount-like fields (Quantity/Tax/Total) are summed across rows; Unit Price
    takes the first value; text fields take the first non-empty value.
    """
    if not col or col not in df.columns:
        return None
    series = df[col]
    if field in _SUM_FIELDS:
        nums = [_to_num(v) for v in series]
        nums = [x for x in nums if x is not None]
        return sum(nums) if nums else None
    if field == "Unit Price":
        for v in series:
            x = _to_num(v)
            if x is not None:
                return x
        return None
    for v in series:  # text: first non-empty
        s = str(v).strip()
        if s and s.lower() != "nan":
            return s
    return None


def _compare_field(field, po_val, inv_val) -> str:
    if po_val is None or inv_val is None:
        return "Missing"
    if field in _NUMERIC_FIELDS:
        a, b = _to_num(po_val), _to_num(inv_val)
        if a is None or b is None:
            return "Missing"
        return "Match" if abs(a - b) <= max(0.01, abs(a) * 0.005) else "Mismatch"
    if field == "GST Number":
        norm = lambda s: re.sub(r"\s+", "", str(s)).upper()
        return "Match" if norm(po_val) == norm(inv_val) else "Mismatch"
    return "Match" if _norm_text(po_val) == _norm_text(inv_val) else "Mismatch"


def _rating(pct: float) -> str:
    if pct >= 95:
        return "Excellent"
    if pct >= 80:
        return "Good"
    if pct >= 60:
        return "Review Required"
    return "High Risk"


def _disp(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def validate_po_invoice(po_df, inv_df, po_map: dict, inv_map: dict) -> dict:
    """Validate an invoice against a PO field by field. *_map: field -> column."""
    results = []
    matched = mismatched = missing = 0
    for field in PO_FIELDS:
        po_val = _extract_field(po_df, po_map.get(field), field)
        inv_val = _extract_field(inv_df, inv_map.get(field), field)
        status = _compare_field(field, po_val, inv_val)
        if status == "Match":
            matched += 1
        elif status == "Mismatch":
            mismatched += 1
        else:
            missing += 1
        results.append({
            "Field": field,
            "PO Value": _disp(po_val),
            "Invoice Value": _disp(inv_val),
            "Status": status,
        })

    total = len(PO_FIELDS)
    pct = matched / total * 100 if total else 0.0
    return {
        "results": results,
        "total": total,
        "matched": matched,
        "mismatched": mismatched,
        "missing": missing,
        "match_pct": pct,
        "score": pct,
        "rating": _rating(pct),
    }


def build_validation_excel(result: dict) -> bytes:
    """Excel report: Summary / Detailed Validation / Exceptions (colour-coded)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HEAD = PatternFill("solid", fgColor="1F2937")
    fill_for = {"Match": GREEN, "Mismatch": RED, "Missing": YELLOW}

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Total fields checked", result["total"]])
    ws.append(["Matched", result["matched"]])
    ws.append(["Mismatched", result["mismatched"]])
    ws.append(["Missing", result["missing"]])
    ws.append(["Overall match %", round(result["match_pct"], 1)])
    ws.append(["Compliance score", round(result["score"], 1)])
    ws.append(["Rating", result["rating"]])

    # Sheet 2 — Detailed Validation (all fields, colour-coded status)
    ws2 = wb.create_sheet("Detailed Validation")
    ws2.append(["Field", "PO Value", "Invoice Value", "Status"])
    for r in result["results"]:
        ws2.append([r["Field"], r["PO Value"], r["Invoice Value"], r["Status"]])
        ws2.cell(row=ws2.max_row, column=4).fill = fill_for.get(r["Status"], YELLOW)

    # Sheet 3 — Exceptions (mismatches + missing only)
    ws3 = wb.create_sheet("Exceptions")
    ws3.append(["Field", "PO Value", "Invoice Value", "Status"])
    for r in result["results"]:
        if r["Status"] != "Match":
            ws3.append([r["Field"], r["PO Value"], r["Invoice Value"], r["Status"]])
            ws3.cell(row=ws3.max_row, column=4).fill = fill_for.get(r["Status"], YELLOW)

    for sheet in (ws, ws2, ws3):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEAD

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
