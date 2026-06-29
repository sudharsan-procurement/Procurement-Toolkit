"""
OCR — reading text out of SCANNED documents and images.

A "scanned" PDF (or a photo/screenshot) has no real text inside — it's just a
picture of words. OCR ("optical character recognition") looks at the picture and
works out what the words are.

Engine choice: we use **Tesseract**, the long-standing free open-source OCR
engine. It's light enough to run on the free Streamlit host (heavier engines like
PaddleOCR tend to exceed the free memory limit). The Tesseract program itself is
installed on the server via `packages.txt`; this file is just the Python side.

How a scanned PDF is read:
    each page is rendered to an image with PyMuPDF, then handed to Tesseract.
We render in greyscale at a moderate resolution — enough for accuracy, light on
memory.
"""

from __future__ import annotations

import io

# Resolution for rendering PDF pages before OCR. ~200 DPI is a good accuracy /
# memory trade-off on a small free server.
_OCR_DPI = 200


def tesseract_available() -> bool:
    """True if the Tesseract engine is installed and callable on this machine."""
    try:
        import pytesseract

        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def _reconstruct(data: dict) -> tuple[str, float]:
    """
    Turn Tesseract's word-level data (from image_to_data) into:
        - the page text, with line breaks preserved, and
        - the page's average confidence (0-100) over real words.

    Using image_to_data gives us BOTH text and confidence in a single OCR pass,
    so we get confidence "for free" without slowing OCR down.
    """
    n = len(data["text"])
    lines: dict[tuple, list[str]] = {}
    confs: list[float] = []
    for i in range(n):
        word = (data["text"][i] or "").strip()
        try:
            conf = float(data["conf"][i])
        except (TypeError, ValueError):
            conf = -1.0
        if word and conf >= 0:
            confs.append(conf)
        if word:
            key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
            lines.setdefault(key, []).append(word)

    text = "\n".join(" ".join(words) for _, words in sorted(lines.items()))
    avg_conf = sum(confs) / len(confs) if confs else 0.0
    return text, avg_conf


def ocr_image_with_confidence(image_bytes: bytes) -> tuple[str, list[dict]]:
    """Read text from one image, returning (text, [{"page": 1, "confidence": x}])."""
    import pytesseract
    from pytesseract import Output
    from PIL import Image

    image = Image.open(io.BytesIO(image_bytes))
    data = pytesseract.image_to_data(image, output_type=Output.DICT)
    text, conf = _reconstruct(data)
    return text.strip(), [{"page": 1, "confidence": round(conf, 1)}]


def ocr_pdf_with_confidence(
    pdf_bytes: bytes, dpi: int = _OCR_DPI
) -> tuple[str, list[dict]]:
    """Read a scanned PDF, returning (full_text, per-page confidence reports)."""
    import fitz  # PyMuPDF
    import pytesseract
    from pytesseract import Output
    from PIL import Image

    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)

    page_texts: list[str] = []
    reports: list[dict] = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for idx, page in enumerate(doc, start=1):
            # Render the page to a greyscale image (lighter on memory than colour).
            pix = page.get_pixmap(matrix=matrix, colorspace=fitz.csGRAY)
            image = Image.frombytes("L", (pix.width, pix.height), pix.samples)
            data = pytesseract.image_to_data(image, output_type=Output.DICT)
            text, conf = _reconstruct(data)
            page_texts.append(text)
            reports.append({"page": idx, "confidence": round(conf, 1)})

    return "\n".join(page_texts).strip(), reports


# --- Backwards-compatible wrappers (text only) ---------------------------------
def ocr_image_bytes(image_bytes: bytes) -> str:
    """Read text from a single image (PNG/JPG/TIFF/…)."""
    text, _ = ocr_image_with_confidence(image_bytes)
    return text


def ocr_pdf_bytes(pdf_bytes: bytes, dpi: int = _OCR_DPI) -> str:
    """Read text from a scanned PDF, one page at a time, and join it together."""
    text, _ = ocr_pdf_with_confidence(pdf_bytes, dpi)
    return text


# --- Confidence classification, summary, and Excel export ----------------------
def classify_confidence(conf: float) -> str:
    """Map a 0-100 confidence score to a plain-English quality status."""
    if conf >= 90:
        return "Excellent"
    if conf >= 75:
        return "Good"
    if conf >= 60:
        return "Review Recommended"
    return "Manual Verification Required"


def summarize_confidence(pages: list[dict]) -> dict:
    """Return total pages, average confidence, and the lowest/highest pages."""
    if not pages:
        return {"total_pages": 0, "average": 0.0, "lowest": None, "highest": None}
    lowest = min(pages, key=lambda p: p["confidence"])
    highest = max(pages, key=lambda p: p["confidence"])
    average = sum(p["confidence"] for p in pages) / len(pages)
    return {
        "total_pages": len(pages),
        "average": average,
        "lowest": lowest,
        "highest": highest,
    }


def build_ocr_report_excel(pages: list[dict]) -> bytes:
    """Build an Excel report: Page Number / Confidence Score / Status."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Confidence"

    headers = ["Page Number", "Confidence Score", "Status"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F2937")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for p in pages:
        ws.append([p["page"], round(p["confidence"], 1),
                   classify_confidence(p["confidence"])])

    for col, width in {"A": 14, "B": 18, "C": 28}.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
